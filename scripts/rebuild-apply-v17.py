from __future__ import annotations

from pathlib import Path
import json
import re
import subprocess

ROOT = Path(__file__).resolve().parents[1]
APPLY = ROOT / "apply"
PDF = APPLY / "pdf"
SCRIPTS = ROOT / "scripts"


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.strip() + "\n", encoding="utf-8")


CONFIG = r'''(() => {
  'use strict';
  window.CITRONEX_SIMPLE_CONFIG = Object.freeze({
    version: '17.0.0',
    brand: 'Citronex / PPO Siechnice',
    privacyUrl: 'https://pposiechnice.pl/?lang=en&page_id=981',
    pdfGeneratorUrl: './pdf/',
    timeZone: 'Europe/Warsaw',
    storageKey: 'citronex_simple_application_v17',
    draftMaxAgeMs: 24 * 60 * 60 * 1000,
    maxMailtoLength: 7600,
    recruiters: Object.freeze([
      Object.freeze({ id: 'yana', name: 'Yana Radushynska', email: 'yana.radushynska@pposiechnice.pl', phone: '+48 797 066 987', initials: 'YR' }),
      Object.freeze({ id: 'yuliia', name: 'Yuliia Korniienko', email: 'yuliia.korniienko@pposiechnice.pl', phone: '+48 506 845 667', initials: 'YK' }),
      Object.freeze({ id: 'fariz', name: 'Fariz Injaev', email: 'fariz.injaev@pposiechnice.pl', phone: '+48 504 165 739', initials: 'FI' }),
      Object.freeze({ id: 'oleksandr', name: 'Oleksandr Kiris', email: 'oleksandr.kiris@pposiechnice.pl', phone: '+48 502 251 384', initials: 'OK' }),
      Object.freeze({ id: 'maksym', name: 'Maksym Saliuk', email: 'maksym.saliuk@pposiechnice.pl', phone: '+48 506 845 637', initials: 'MS' }),
      Object.freeze({ id: 'anastasiia', name: 'Anastasiia Derepa', email: 'anastasiia.derepa@citronex.pl', phone: '+48 797 684 159', initials: 'AD' })
    ]),
    locations: Object.freeze(['siechnice', 'ryczywol', 'bogatynia', 'zgorzelec', 'pruszcz', 'any']),
    queryParams: Object.freeze({
      language: 'lang', recruiter: 'recruiter', source: 'src', campaign: 'campaign', vacancy: 'vacancy',
      location: 'location', partner: 'partner', group: 'group'
    }),
    excelColumns: Object.freeze([
      'ID zgłoszenia', 'Data zgłoszenia', 'SLA do', 'Język', 'Rekruter', 'E-mail rekrutera',
      'Ankietę wypełnia', 'Osoba / partner wypełniający', 'Kod grupy / partnera',
      'Preferowana lokalizacja', 'Imię', 'Nazwisko', 'Telefon', 'Komunikator', 'E-mail kandydata',
      'Obywatelstwo', 'Kraj pobytu', 'Miasto', 'Wiek', 'W Polsce', 'Dokument pobytowy',
      'Stanowisko', 'Doświadczenie', 'Gotowość', 'Praca zmianowa', 'Zakwaterowanie',
      'Źródło deklarowane', 'Szczegóły źródła / polecający', 'Źródło linku', 'Kampania', 'Wakacja / oferta',
      'Komentarz', 'Status', 'Data pierwszego kontaktu', 'Liczba prób kontaktu', 'Następny kontakt',
      'Wynik rozmowy', 'Decyzja', 'Powód odmowy', 'Uwagi rekrutera'
    ])
  });
})();'''

INDEX = r'''<!doctype html>
<html lang="pl" dir="ltr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
  <meta name="theme-color" content="#075b39">
  <meta name="description" content="Prosty wielojęzyczny formularz rekrutacyjny Citronex / PPO Siechnice">
  <meta name="referrer" content="no-referrer">
  <meta name="format-detection" content="telephone=no">
  <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
  <meta http-equiv="Pragma" content="no-cache">
  <meta http-equiv="Expires" content="0">
  <meta http-equiv="Content-Security-Policy" content="default-src 'self'; script-src 'self'; style-src 'self'; img-src 'self' data:; connect-src 'none'; object-src 'none'; base-uri 'self'; form-action 'none'; frame-ancestors 'none';">
  <title>Citronex / PPO Siechnice — Rekrutacja</title>
  <link rel="icon" href="../assets/icon-192.png">
  <link rel="stylesheet" href="styles.css?v=17.0.0">
  <script defer src="config.js?v=17.0.0"></script>
  <script defer src="translations.js?v=17.0.0"></script>
  <script defer src="app.js?v=17.0.0"></script>
</head>
<body>
  <noscript><div class="noscript">Ta strona wymaga włączonego JavaScript.</div></noscript>
  <header class="site-header">
    <div class="header-inner">
      <a class="brand" href="./" aria-label="Citronex / PPO Siechnice — formularz rekrutacyjny">
        <span class="logo-frame"><img src="../assets/citronex-logo.jpg" alt="Citronex" width="138" height="52"></span>
        <span class="brand-copy"><strong>Citronex / PPO Siechnice</strong><small id="headerSubtitle">Formularz kandydata</small></span>
      </a>
      <span class="header-note"><span class="header-note-dot" aria-hidden="true"></span><span id="headerNote">Pierwszy kontakt do 24 godzin</span></span>
    </div>
  </header>
  <main id="app" class="app-shell" aria-live="polite"></main>
  <footer class="site-footer">
    <p>Formularz nie zapisuje danych na GitHub. <a id="footerPdf" href="pdf/">Generator PDF</a> · <a href="https://pposiechnice.pl/?lang=en&page_id=981" target="_blank" rel="noopener noreferrer">Prywatność</a></p>
  </footer>
</body>
</html>'''

STYLES = r''':root{
  color-scheme:light;--page:#eef5f1;--surface:#fff;--surface-soft:#f7faf8;--ink:#13251b;--muted:#64756b;
  --line:#d5e2db;--line-strong:#b9cec3;--green:#075b39;--green-dark:#043d28;--green-light:#15935a;
  --green-soft:#e9f7ef;--gold:#e5a62d;--gold-soft:#fff7e6;--danger:#b42318;--danger-soft:#fff1f0;
  --blue:#175cd3;--blue-soft:#eff7ff;--shadow:0 22px 60px rgba(6,61,39,.11)
}
*{box-sizing:border-box}
html{min-width:320px;background:var(--page);color:var(--ink);font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;-webkit-text-size-adjust:100%;scroll-behavior:smooth}
body{min-width:320px;min-height:100vh;margin:0;background:radial-gradient(circle at 0 0,rgba(21,147,90,.10),transparent 30rem),radial-gradient(circle at 100% 12%,rgba(229,166,45,.08),transparent 27rem),linear-gradient(180deg,#f8fbf9 0,var(--page) 52%,#eaf2ed 100%)}
button,input,select,textarea{font:inherit}button,a{-webkit-tap-highlight-color:transparent}button{cursor:pointer}a{color:var(--green)}
.hidden{display:none!important}.invisible{visibility:hidden}.ltr{direction:ltr;unicode-bidi:isolate}.sr-only{position:absolute!important;width:1px!important;height:1px!important;padding:0!important;margin:-1px!important;overflow:hidden!important;clip:rect(0,0,0,0)!important;white-space:nowrap!important;border:0!important}
.noscript{padding:16px;background:var(--danger-soft);color:var(--danger);font-weight:800;text-align:center}
.site-header{position:relative;z-index:1;border-bottom:1px solid rgba(7,91,57,.10);background:#fff;box-shadow:0 4px 18px rgba(5,55,35,.045)}
.header-inner{width:min(780px,100%);min-height:66px;margin:auto;padding:8px 14px;display:flex;align-items:center;justify-content:space-between;gap:14px}
.brand{min-width:0;display:flex;align-items:center;gap:11px;color:var(--ink);text-decoration:none}.logo-frame{width:82px;height:47px;flex:none;display:grid;place-items:center;overflow:hidden;border:1px solid var(--line);border-radius:14px;background:#fff;box-shadow:0 7px 19px rgba(5,72,44,.10)}.logo-frame img{width:100%;height:100%;object-fit:contain;padding:3px}.brand-copy{min-width:0}.brand-copy strong{display:block;font-size:14px;line-height:1.2;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.brand-copy small{display:block;margin-top:3px;color:var(--muted);font-size:10px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.header-note{max-width:235px;display:flex;align-items:center;gap:7px;color:var(--muted);font-size:10px;line-height:1.35;text-align:right}.header-note-dot{width:8px;height:8px;flex:none;border-radius:50%;background:var(--green-light);box-shadow:0 0 0 4px rgba(21,147,90,.10)}
.app-shell{width:min(780px,100%);min-height:calc(100vh - 135px);margin:auto;padding:18px 9px 70px}.panel{position:relative;overflow:hidden;border:1px solid rgba(7,91,57,.10);border-radius:26px;background:#fff;box-shadow:var(--shadow);isolation:isolate}.panel::before{content:"";position:absolute;inset:0 0 auto;height:4px;background:linear-gradient(90deg,var(--green-dark),var(--green-light) 72%,var(--gold))}.opening-panel,.form-panel,.review-panel{padding:25px 24px}
.opening-copy{text-align:center;padding:7px 0 19px}.opening-copy.compact{display:grid;grid-template-columns:58px minmax(0,1fr);grid-template-rows:auto auto;align-items:center;column-gap:14px;text-align:left}.opening-copy.compact .opening-icon{grid-row:1/3;margin:0}.opening-copy.compact h1,.opening-copy.compact p{grid-column:2}.opening-copy.compact p{margin:5px 0 0}.eyebrow{display:inline-flex;margin-bottom:12px;padding:6px 9px;border:1px solid #b9ddca;border-radius:999px;background:#eff9f4;color:var(--green);font-size:9px;font-weight:900;letter-spacing:.10em}.opening-icon{width:62px;height:62px;margin:0 auto 13px;display:grid;place-items:center;border:1px solid #d5e9df;border-radius:21px;background:linear-gradient(145deg,var(--green-soft),var(--gold-soft));box-shadow:0 11px 27px rgba(20,73,47,.08);font-size:30px}h1,h2,h3,p{overflow-wrap:anywhere}.opening-copy h1,.step-heading h1,.review-heading h1{margin:0;color:#102d20;font-size:clamp(28px,7vw,40px);line-height:1.08;letter-spacing:-.035em}.opening-copy p,.step-heading p,.review-heading p{max-width:610px;margin:9px auto 0;color:var(--muted);font-size:14px;line-height:1.52}
.search-input{width:100%;min-height:56px;margin-bottom:13px;padding:0 17px;border:1.5px solid var(--line-strong);border-radius:16px;background:#fff;color:var(--ink);font-size:16px}.search-input:focus,.field input:focus,.field select:focus,.field textarea:focus{outline:0;border-color:var(--green-light);box-shadow:0 0 0 4px rgba(21,147,90,.11)}.language-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:9px}.language-card{position:relative;min-width:0;min-height:84px;display:grid;grid-template-columns:auto minmax(0,1fr);align-items:center;gap:10px;padding:12px 11px;border:1px solid var(--line);border-radius:18px;background:linear-gradient(180deg,#fff,#fbfdfc);color:var(--ink);text-align:left;box-shadow:0 5px 15px rgba(7,84,54,.035)}.language-card:hover,.language-card:focus-visible,.recruiter-card:hover,.recruiter-card:focus-visible{outline:0;border-color:var(--green-light);box-shadow:0 10px 23px rgba(7,84,54,.09)}.language-card.suggested-card,.recruiter-card.suggested-card{border-color:#9ccfb4;background:linear-gradient(145deg,#effaf4,#fff)}.language-flag{font-size:28px}.language-name{min-width:0}.language-name strong{display:block;font-size:13px;line-height:1.25;overflow-wrap:anywhere}.language-name small{display:block;margin-top:3px;color:var(--muted);font-size:9px}.language-code{position:absolute;right:8px;bottom:7px;min-width:25px;padding:3px 5px;border-radius:8px;background:#eff4f1;color:#718278;font-size:8px;font-weight:900;text-align:center}.more-button{width:100%;min-height:48px;margin-top:11px;border:1px solid var(--line);border-radius:15px;background:var(--surface-soft);color:var(--green);font-weight:800}
.draft-box{margin:0 0 13px;padding:13px 14px;border:1px solid #bdd7ff;border-radius:16px;background:var(--blue-soft);color:#174a9c}.draft-box strong{display:block;margin-bottom:10px;font-size:13px}.inline-actions{display:flex;gap:8px;flex-wrap:wrap}.back-top{margin:0 0 9px}.info-strip{margin:0 0 14px;padding:11px 13px;border:1px solid #bed8c9;border-radius:14px;background:#eff9f4;color:#24563e;font-size:11px;font-weight:700;line-height:1.45}.recruiter-list{display:grid;gap:9px}.recruiter-card{min-width:0;min-height:92px;display:grid;grid-template-columns:55px minmax(0,1fr) 24px;align-items:center;gap:12px;padding:13px 14px;border:1px solid var(--line);border-radius:19px;background:linear-gradient(180deg,#fff,#fbfdfc);color:var(--ink);text-align:left}.avatar{width:55px;height:55px;display:grid;place-items:center;border-radius:18px;background:linear-gradient(145deg,var(--green-soft),#d6f0e2);color:var(--green);font-size:15px;font-weight:900}.recruiter-copy{min-width:0}.recruiter-copy strong{display:block;font-size:15px}.recruiter-copy small{display:block;margin-top:4px;color:var(--muted);font-size:10px;line-height:1.3;overflow-wrap:anywhere}.recruiter-copy .phone{color:#315d48;font-size:11px;font-weight:800}.card-arrow{color:var(--green-light);font-size:29px}
.form-topbar{display:flex;align-items:center;justify-content:space-between;gap:10px;margin-bottom:12px;flex-wrap:wrap}.language-pill{display:inline-flex;align-items:center;min-height:33px;padding:7px 10px;border:1px solid var(--line);border-radius:11px;background:#f5f8f6;color:var(--muted);font-size:10px;font-weight:800}.form-topbar>div{display:flex;gap:3px;flex-wrap:wrap}.text-button{min-height:40px;padding:7px 9px;border:0;border-radius:10px;background:transparent;color:var(--green);font-weight:800}.text-button:hover,.text-button:focus-visible{outline:0;background:var(--green-soft)}.selected-recruiter{display:grid;grid-template-columns:46px minmax(0,1fr);align-items:center;gap:11px;margin-bottom:14px;padding:12px 13px;border:1px solid #acd5bf;border-radius:18px;background:linear-gradient(135deg,#eefaf4,#fff)}.compact-avatar{width:46px;height:46px;border-radius:15px}.selected-copy{min-width:0}.selected-copy small{display:block;color:var(--muted);font-size:9px;font-weight:900;text-transform:uppercase}.selected-copy strong{display:block;margin-top:2px;font-size:13px}.selected-copy span{display:block;margin-top:3px;color:#315d48;font-size:10px;font-weight:800}
.progress-box{margin:0 0 17px;padding:11px 13px;border:1px solid #dce9e2;border-radius:16px;background:#f8fbf9}.progress-meta{display:flex;align-items:center;justify-content:space-between;gap:8px;margin-bottom:7px}.progress-meta strong{color:var(--green);font-size:11px}.progress-meta span{color:var(--muted);font-size:10px;text-align:right}.progress-track{height:7px;overflow:hidden;border-radius:999px;background:#dfe8e3}.progress-fill{display:block;height:100%;border-radius:inherit;background:linear-gradient(90deg,var(--green-dark),var(--green-light))}.progress-1{width:25%}.progress-2{width:50%}.progress-3{width:75%}.progress-4{width:100%}.step-heading{display:grid;grid-template-columns:38px minmax(0,1fr);align-items:start;gap:11px;margin-bottom:18px}.step-number{width:38px;height:38px;display:grid;place-items:center;border-radius:13px;background:#dff3e8;color:var(--green);font-size:14px;font-weight:900}.step-heading h1{font-size:clamp(25px,6vw,33px)}.step-heading p{max-width:none;margin:7px 0 0;font-size:13px}
.form-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:0 13px}.field{min-width:0;margin-bottom:16px}.field.full{grid-column:1/-1}.field label,.field-label{display:block;margin-bottom:8px;color:#1c392a;font-size:13px;font-weight:850}.required{color:var(--danger)}.optional{color:var(--muted);font-size:10px;font-weight:600}.field input,.field select,.field textarea{width:100%;min-height:56px;padding:13px 14px;border:1.5px solid #c9d9d1;border-radius:15px;background:#fff;color:var(--ink);font-size:16px}.field textarea{min-height:112px;line-height:1.5;resize:vertical}.field input.invalid,.field select.invalid,.field textarea.invalid{border-color:var(--danger);background:var(--danger-soft)}.field-hint{display:block;margin-top:7px;color:var(--muted);font-size:10px;line-height:1.42}.field-error{margin-top:6px;color:var(--danger);font-size:10px;font-weight:800;line-height:1.35}.partner-box{margin:0 0 16px;padding:14px;border:1px solid #b8ddc8;border-radius:19px;background:linear-gradient(145deg,#f1faf5,#fff)}.partner-box .field:last-child{margin-bottom:0}
.choice-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:9px}.choice-grid.one{grid-template-columns:1fr}.choice-grid.three{grid-template-columns:repeat(3,minmax(0,1fr))}.choice-card{position:relative;min-width:0;min-height:72px;display:grid;grid-template-columns:42px minmax(0,1fr) 23px;align-items:center;gap:9px;padding:11px 12px;border:1px solid var(--line);border-radius:17px;background:linear-gradient(180deg,#fff,#fbfdfc);color:var(--ink);cursor:pointer}.choice-card input{position:absolute;opacity:0;pointer-events:none}.choice-card.selected{border-color:var(--green-light);background:#effaf4;box-shadow:0 0 0 3px rgba(21,147,90,.10)}.choice-icon{width:42px;height:42px;display:grid;place-items:center;border-radius:14px;background:#edf7f2;color:var(--green);font-size:16px;font-weight:900}.choice-card>span:nth-child(3){min-width:0;font-size:12px;font-weight:800;line-height:1.3;overflow-wrap:anywhere}.choice-check{width:22px;height:22px;display:grid;place-items:center;border:1px solid var(--line-strong);border-radius:50%;background:#fff;color:transparent;font-size:11px;font-weight:900}.choice-card.selected .choice-check{border-color:var(--green);background:var(--green);color:#fff}.consent-box{display:flex!important;align-items:flex-start;gap:10px;margin:0!important;padding:13px;border:1px solid var(--line);border-radius:15px;background:var(--surface-soft);font-size:11px!important;font-weight:600!important;line-height:1.5}.consent-box input{width:22px!important;height:22px!important;min-height:22px!important;flex:none;margin:1px 0 0;padding:0!important}
.sticky-actions{position:relative;display:flex;align-items:center;justify-content:space-between;gap:9px;margin:24px 0 0;padding:14px 0 0;border-top:1px solid var(--line);background:#fff}.button{min-height:54px;padding:0 18px;border:1px solid var(--line);border-radius:15px;background:#fff;color:var(--ink);font-weight:850}.button:hover:not(:disabled){box-shadow:0 8px 20px rgba(23,66,46,.08)}.button.primary{border-color:var(--green);background:var(--green);color:#fff}.button.primary:hover{background:var(--green-dark)}.button.secondary{background:var(--surface-soft)}.button.small{min-height:40px;padding:0 12px;font-size:11px}.button:disabled{cursor:not-allowed;opacity:.58}.sticky-actions .primary{flex:1}
.review-heading{display:grid;grid-template-columns:58px minmax(0,1fr);align-items:center;gap:13px;margin-bottom:14px}.review-heading h1{font-size:clamp(27px,6vw,35px)}.review-heading p{margin:5px 0 0}.success-icon{width:58px;height:58px;display:grid;place-items:center;border-radius:19px;background:var(--green-soft);color:var(--green);font-size:27px;font-weight:900}.recipient-card{display:grid;grid-template-columns:50px minmax(0,1fr);align-items:center;gap:12px;margin-bottom:14px;padding:14px 15px;border-radius:19px;background:linear-gradient(135deg,var(--green-dark),#08764a);color:#fff;box-shadow:0 13px 28px rgba(7,92,57,.18)}.recipient-card .avatar{width:50px;height:50px;background:rgba(255,255,255,.16);color:#fff}.recipient-card small,.recipient-card strong,.recipient-card em{display:block}.recipient-card small{color:rgba(255,255,255,.72);font-size:9px;text-transform:uppercase}.recipient-card strong{margin-top:2px;font-size:14px}.recipient-card em{margin-top:4px;color:rgba(255,255,255,.88);font-size:10px;font-style:normal;overflow-wrap:anywhere}.review-sections{display:grid;gap:10px}.review-section{overflow:hidden;border:1px solid var(--line);border-radius:18px;background:#fff}.review-section h2{margin:0;padding:12px 14px;border-bottom:1px solid var(--line);background:#f5f9f7;font-size:14px}.review-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr))}.review-item{min-width:0;padding:11px 14px;border-bottom:1px solid #e8efeb}.review-item:nth-child(odd){border-right:1px solid #e8efeb}.review-item small{display:block;margin-bottom:4px;color:var(--muted);font-size:9px;font-weight:800;text-transform:uppercase}.review-item strong{display:block;font-size:12px;line-height:1.4;overflow-wrap:anywhere;white-space:pre-wrap}.send-box{display:grid;grid-template-columns:48px minmax(0,1fr);gap:11px;margin-top:16px;padding:16px;border:1px solid #a9d7bd;border-radius:20px;background:linear-gradient(145deg,#ecf9f2,#fff)}.send-icon{width:48px;height:48px;display:grid;place-items:center;border-radius:15px;background:var(--green);color:#fff;font-size:23px}.send-box h2{margin:2px 0 4px;color:var(--green-dark);font-size:18px}.send-box p{margin:0;color:var(--muted);font-size:11px;line-height:1.45}.send-button{grid-column:1/-1;width:100%}.send-result{grid-column:1/-1;padding:11px 12px;border:1px solid #a7dfc0;border-radius:13px;background:#ecfdf3;color:#067647;font-size:11px;font-weight:700}.pdf-card{display:grid;grid-template-columns:48px minmax(0,1fr) auto;align-items:center;gap:11px;margin-top:13px;padding:14px;border:1px solid #efcf8c;border-radius:18px;background:var(--gold-soft);color:var(--ink);text-decoration:none}.pdf-icon{width:48px;height:48px;display:grid;place-items:center;border-radius:14px;background:#fff;color:#a66100;font-size:12px;font-weight:900}.pdf-card strong,.pdf-card small{display:block}.pdf-card strong{font-size:13px}.pdf-card small{margin-top:4px;color:var(--muted);font-size:10px;line-height:1.4}.pdf-card b{color:#8a4b00;font-size:10px}.review-actions{display:flex;align-items:center;justify-content:space-between;gap:9px;margin-top:15px;flex-wrap:wrap}.site-footer{padding:0 15px 27px;text-align:center;color:var(--muted);font-size:10px;line-height:1.5}
html[dir="rtl"] .opening-copy.compact,html[dir="rtl"] .recruiter-card,html[dir="rtl"] .field,html[dir="rtl"] .review-item{text-align:right}html[dir="rtl"] .language-code{right:auto;left:8px}
@media(max-width:700px){.header-note{display:none}.opening-panel,.form-panel,.review-panel{padding:19px 15px}.form-grid{grid-template-columns:1fr}.field.full{grid-column:auto}.language-grid{grid-template-columns:repeat(2,minmax(0,1fr))}.choice-grid.three{grid-template-columns:repeat(2,minmax(0,1fr))}.review-grid{grid-template-columns:1fr}.review-item:nth-child(odd){border-right:0}.sticky-actions{display:grid;grid-template-columns:1fr 1fr}.sticky-actions .primary{width:100%}}
@media(max-width:440px){.header-inner{padding:8px 10px}.logo-frame{width:72px;height:42px}.brand-copy strong{max-width:210px;font-size:12px}.brand-copy small{font-size:9px}.app-shell{padding:10px 6px 52px}.panel{border-radius:19px}.opening-panel,.form-panel,.review-panel{padding:17px 12px}.opening-copy.compact{grid-template-columns:49px minmax(0,1fr);column-gap:10px}.opening-icon{width:54px;height:54px;border-radius:18px;font-size:26px}.opening-copy.compact .opening-icon{width:49px;height:49px}.opening-copy h1,.step-heading h1,.review-heading h1{font-size:27px}.language-card{min-height:78px;padding:10px 9px}.language-flag{font-size:25px}.recruiter-card{grid-template-columns:48px minmax(0,1fr) 18px;gap:9px;padding:11px}.avatar{width:48px;height:48px;border-radius:15px}.choice-grid,.choice-grid.three{grid-template-columns:1fr}.choice-card{min-height:66px}.step-heading{grid-template-columns:34px minmax(0,1fr)}.step-number{width:34px;height:34px}.review-heading{grid-template-columns:50px minmax(0,1fr)}.success-icon{width:50px;height:50px}.pdf-card{grid-template-columns:44px minmax(0,1fr)}.pdf-card b{grid-column:2}.sticky-actions{grid-template-columns:1fr}.sticky-actions .invisible{display:none}.button{width:100%}.review-actions{display:grid;grid-template-columns:1fr}}
@media(max-width:350px){.language-grid{grid-template-columns:1fr}.brand-copy strong{max-width:175px}.recruiter-copy small{font-size:9px}}
@media(prefers-reduced-motion:reduce){html{scroll-behavior:auto}}
'''

PDF_INDEX = r'''<!doctype html>
<html lang="pl" dir="ltr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
  <meta name="theme-color" content="#075b39">
  <meta name="description" content="Opcjonalny generator PDF kandydata — bez wysyłania danych na serwer">
  <meta name="referrer" content="no-referrer">
  <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
  <meta http-equiv="Content-Security-Policy" content="default-src 'self'; script-src 'self'; style-src 'self'; img-src 'self' blob: data:; connect-src 'none'; object-src 'none'; base-uri 'self'; form-action 'none'; frame-ancestors 'none';">
  <title>Generator PDF — Citronex / PPO Siechnice</title>
  <link rel="icon" href="../../assets/icon-192.png">
  <link rel="stylesheet" href="styles.css?v=2.0.0">
  <script defer src="../translations.js?v=17.0.0"></script>
  <script defer src="app.js?v=2.0.0"></script>
</head>
<body>
  <header class="site-header no-print">
    <div class="header-inner">
      <a class="brand" href="../"><span class="logo-frame"><img src="../../assets/citronex-logo.jpg" alt="Citronex" width="138" height="52"></span><span><strong>Citronex / PPO Siechnice</strong><small id="brandSubtitle">Generator PDF</small></span></a>
      <a class="back-link" href="../" id="backLink">← Ankieta</a>
    </div>
  </header>
  <main id="pdfApp" class="page-shell" aria-live="polite"></main>
  <footer class="site-footer no-print"><p id="footerText">Dane i zdjęcia pozostają na tym urządzeniu. Nic nie jest wysyłane na serwer.</p></footer>
</body>
</html>'''

PDF_STYLES = r''':root{color-scheme:light;--page:#eef5f1;--surface:#fff;--ink:#14231b;--muted:#64736b;--line:#d6e3dc;--green:#075b39;--green-2:#0a7a4b;--green-soft:#eaf7f0;--gold:#e8aa32;--danger:#b42318;--shadow:0 20px 55px rgba(7,72,45,.10)}
*{box-sizing:border-box}html{min-width:320px;font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;background:var(--page);color:var(--ink);-webkit-text-size-adjust:100%}body{min-width:320px;min-height:100vh;margin:0;background:radial-gradient(circle at 0 0,rgba(18,148,88,.10),transparent 28rem),radial-gradient(circle at 100% 14%,rgba(232,170,50,.08),transparent 24rem),var(--page)}button,input,select,textarea{font:inherit}button{cursor:pointer}a{color:var(--green-2)}.hidden{display:none!important}.ltr{direction:ltr;unicode-bidi:isolate}.site-header{position:relative;z-index:1;background:#fff;border-bottom:1px solid rgba(7,91,57,.10);box-shadow:0 4px 18px rgba(5,55,35,.045)}.header-inner{width:min(1120px,100%);min-height:64px;margin:auto;padding:8px 14px;display:flex;align-items:center;justify-content:space-between;gap:14px}.brand{min-width:0;display:flex;align-items:center;gap:11px;color:var(--ink);text-decoration:none}.logo-frame{width:82px;height:46px;display:grid;place-items:center;overflow:hidden;border-radius:13px;background:#fff;border:1px solid var(--line)}.logo-frame img{width:100%;height:100%;object-fit:contain;padding:3px}.brand strong,.brand small{display:block}.brand strong{font-size:14px}.brand small{margin-top:3px;color:var(--muted);font-size:10px}.back-link{padding:9px 11px;border-radius:11px;background:var(--green-soft);color:var(--green);font-size:12px;font-weight:800;text-decoration:none}.page-shell{width:min(1120px,100%);margin:auto;padding:18px 12px 70px}.generator-grid{display:grid;grid-template-columns:minmax(0,420px) minmax(0,1fr);gap:18px;align-items:start}.panel{min-width:0;background:#fff;border:1px solid rgba(7,91,57,.10);border-radius:23px;box-shadow:var(--shadow)}.controls{position:relative;padding:22px}.preview-panel{min-width:0;padding:18px;background:#e5ebe7}.eyebrow{display:inline-flex;padding:5px 8px;border-radius:999px;background:var(--green-soft);color:var(--green);font-size:9px;font-weight:900;letter-spacing:.08em;text-transform:uppercase}.controls h1{margin:10px 0 6px;font-size:30px;line-height:1.05;letter-spacing:-.035em}.lead{margin:0 0 16px;color:var(--muted);font-size:13px;line-height:1.5}.privacy-note,.status-note{margin:0 0 16px;padding:11px 12px;border:1px solid #b9dfca;border-radius:14px;background:#f1fbf5;color:#315c48;font-size:11px;line-height:1.45}.status-note.error{border-color:#f3b4ae;background:#fff1f0;color:var(--danger)}.field{min-width:0;margin-bottom:14px}.field label,.field-label{display:block;margin-bottom:7px;font-size:12px;font-weight:850}.field input,.field select,.field textarea{width:100%;min-height:53px;padding:12px 13px;border:1.5px solid #c9d9d1;border-radius:14px;background:#fff;color:var(--ink);font-size:16px}.field textarea{min-height:95px;resize:vertical;line-height:1.45}.field input:focus,.field select:focus,.field textarea:focus{outline:0;border-color:var(--green-2);box-shadow:0 0 0 4px rgba(10,122,75,.11)}.field-grid{display:grid;grid-template-columns:1fr 1fr;gap:0 10px}.upload-box{position:relative;min-height:110px;display:grid;place-items:center;align-content:center;text-align:center;padding:15px;border:2px dashed #8fc3a8;border-radius:18px;background:linear-gradient(145deg,#fff,#f4fbf7)}.upload-box input{position:absolute;inset:0;width:100%;height:100%;opacity:0;cursor:pointer}.upload-box strong{color:var(--green);font-size:14px}.upload-box small{display:block;margin-top:5px;color:var(--muted);font-size:10px}.file-list{display:grid;gap:7px;margin-top:10px}.file-row{min-width:0;display:grid;grid-template-columns:48px minmax(0,1fr) auto;align-items:center;gap:9px;padding:8px;border:1px solid var(--line);border-radius:13px;background:#fff}.file-thumb{width:48px;height:48px;display:grid;place-items:center;overflow:hidden;border-radius:10px;background:#eef3f0;color:var(--muted);font-size:9px;text-align:center}.file-thumb img{width:100%;height:100%;object-fit:cover}.file-copy{min-width:0}.file-copy strong{display:block;font-size:11px;overflow-wrap:anywhere}.file-copy small{display:block;margin-top:2px;color:var(--muted);font-size:9px}.file-actions{display:flex;gap:4px;flex-wrap:wrap;justify-content:flex-end}.file-actions button{width:31px;height:31px;border:1px solid var(--line);border-radius:9px;background:#f6f9f7;color:var(--green);font-weight:900}.file-actions button.danger{color:var(--danger);background:#fff0ef}.actions{display:grid;grid-template-columns:1fr auto;gap:9px;margin-top:17px}.button{min-height:54px;padding:0 17px;border:1px solid var(--line);border-radius:15px;background:#fff;color:var(--ink);font-weight:850}.button.primary{border-color:var(--green);background:linear-gradient(135deg,var(--green),var(--green-2));color:#fff;box-shadow:0 11px 23px rgba(7,91,57,.18)}.button.secondary{background:#f6f9f7}.help{margin:9px 0 0;color:var(--muted);font-size:9px;line-height:1.4;text-align:center}.preview-title{display:flex;align-items:center;justify-content:space-between;gap:10px;margin:0 0 11px;color:#3c5045;font-size:11px;font-weight:900;text-transform:uppercase}.page-count{padding:4px 7px;border-radius:8px;background:#d5dfda;color:#4f6358;font-size:9px}.pdf-pages{display:grid;gap:16px;justify-items:center}.pdf-sheet{width:min(100%,794px);aspect-ratio:210/297;padding:48px;background:#fff;box-shadow:0 12px 35px rgba(29,45,36,.18);overflow:hidden;page-break-after:always}.pdf-cover{display:flex;flex-direction:column}.pdf-brand{display:flex;align-items:center;justify-content:space-between;gap:15px;padding-bottom:18px;border-bottom:3px solid var(--green)}.pdf-brand img{width:150px;height:58px;object-fit:contain}.pdf-brand span{color:var(--green);font-size:13px;font-weight:900;text-transform:uppercase}.pdf-title{margin:42px 0 27px;font-size:34px;line-height:1.12;color:#143323}.pdf-data{display:grid;grid-template-columns:170px 1fr;border-top:1px solid var(--line)}.pdf-data dt,.pdf-data dd{margin:0;padding:12px 10px;border-bottom:1px solid var(--line)}.pdf-data dt{color:#5d6d64;font-size:11px;font-weight:850;text-transform:uppercase}.pdf-data dd{font-size:14px;font-weight:700;overflow-wrap:anywhere}.pdf-note{margin-top:25px;padding:15px;border-left:4px solid var(--gold);background:#fff9ec;white-space:pre-wrap;font-size:13px;line-height:1.55}.pdf-footer{margin-top:auto;padding-top:18px;border-top:1px solid var(--line);display:flex;justify-content:space-between;color:#79877f;font-size:9px}.photo-sheet{display:grid;grid-template-columns:1fr;gap:12px;padding:24px}.photo-sheet.photos-2{grid-template-columns:1fr 1fr}.photo-sheet.photos-4{grid-template-columns:1fr 1fr;grid-template-rows:1fr 1fr}.photo-sheet figure{min-width:0;min-height:0;margin:0;display:grid;grid-template-rows:1fr auto;place-items:center;overflow:hidden}.photo-sheet figure img{max-width:100%;max-height:100%;object-fit:contain}.photo-sheet figcaption{max-width:100%;padding-top:6px;color:#65766c;font-size:9px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}.rot-0{transform:none}.rot-90{transform:rotate(90deg)}.rot-180{transform:rotate(180deg)}.rot-270{transform:rotate(270deg)}.empty-preview{min-height:360px;display:grid;place-items:center;text-align:center;color:#75847c;background:rgba(255,255,255,.5);border:1px dashed #aebdb5;border-radius:18px;padding:30px}.site-footer{padding:0 15px 28px;text-align:center;color:var(--muted);font-size:10px}
@media(max-width:820px){.generator-grid{grid-template-columns:1fr}.preview-panel{padding:10px}.pdf-sheet{padding:30px}.pdf-title{margin-top:30px;font-size:27px}.pdf-data{grid-template-columns:120px 1fr}}
@media(max-width:520px){.page-shell{padding:10px 7px 60px}.panel{border-radius:19px}.controls{padding:17px}.controls h1{font-size:27px}.field-grid{grid-template-columns:1fr}.actions{grid-template-columns:1fr}.button{width:100%}.file-row{grid-template-columns:44px minmax(0,1fr)}.file-actions{grid-column:1/-1;justify-content:flex-start}.preview-panel{overflow:hidden}.pdf-sheet{padding:21px}.pdf-brand img{width:112px;height:45px}.pdf-brand span{font-size:9px}.pdf-title{font-size:23px}.pdf-data{grid-template-columns:94px 1fr}.pdf-data dt,.pdf-data dd{padding:8px 6px}.pdf-data dt{font-size:8px}.pdf-data dd{font-size:10px}.pdf-note{font-size:10px}.brand strong{font-size:12px}.back-link{font-size:10px}.photo-sheet.photos-2,.photo-sheet.photos-4{grid-template-columns:1fr;grid-template-rows:auto}}
@media print{@page{size:A4;margin:0}html,body{background:#fff!important}.no-print,.controls,.preview-title,.empty-preview{display:none!important}.page-shell,.generator-grid,.preview-panel,.pdf-pages{display:block!important;width:auto!important;margin:0!important;padding:0!important;background:#fff!important;border:0!important;box-shadow:none!important}.pdf-sheet{width:210mm!important;height:297mm!important;aspect-ratio:auto!important;margin:0!important;padding:14mm!important;box-shadow:none!important;break-after:page;page-break-after:always;overflow:hidden}.pdf-sheet:last-child{break-after:auto;page-break-after:auto}.photo-sheet{padding:10mm!important}.photo-sheet.photos-2{grid-template-columns:1fr 1fr!important}.photo-sheet.photos-4{grid-template-columns:1fr 1fr!important;grid-template-rows:1fr 1fr!important}}
'''

PDF_APP = r'''(() => {
  'use strict';
  const app = document.getElementById('pdfApp');
  const MASTER = window.CITRONEX_SIMPLE_I18N;
  if (!app || !MASTER) throw new Error('PDF generator dependencies are missing.');
  const UI = {
    en:{brand:'PDF generator',back:'Application',title:'Create candidate PDF',lead:'Add candidate data and up to 30 document photos. Everything stays on this device.',privacy:'No file is uploaded. Use Print and choose Save as PDF.',language:'Language',firstName:'First name',lastName:'Last name',phone:'Phone',citizenship:'Citizenship',location:'Work location',recruiter:'Recruiter',notes:'Notes',photos:'Document photos',upload:'Add photos',uploadHint:'JPG, PNG, WEBP or HEIC · up to 30 files',layout:'Photos per page',one:'1 photo',two:'2 photos',four:'4 photos',print:'Print / Save as PDF',clear:'Clear',preview:'PDF preview',pages:'pages',empty:'Add candidate data or photos to see the preview.',remove:'Remove',footer:'Data and photos stay on this device. Nothing is sent to a server.',cover:'Candidate card',candidate:'Candidate',created:'Created',noPhotos:'No photos selected',tooMany:'Maximum 30 photos.',badFile:'Only image files are allowed.',tooLarge:'One file is too large. Maximum 15 MB.',totalLarge:'Selected files exceed 100 MB.'},
    pl:{brand:'Generator PDF',back:'Ankieta',title:'Utwórz PDF kandydata',lead:'Dodaj dane kandydata i maksymalnie 30 zdjęć dokumentów. Wszystko pozostaje na tym urządzeniu.',privacy:'Żaden plik nie jest wysyłany. Wybierz Drukuj, a następnie Zapisz jako PDF.',language:'Język',firstName:'Imię',lastName:'Nazwisko',phone:'Telefon',citizenship:'Obywatelstwo',location:'Lokalizacja pracy',recruiter:'Rekruter',notes:'Uwagi',photos:'Zdjęcia dokumentów',upload:'Dodaj zdjęcia',uploadHint:'JPG, PNG, WEBP lub HEIC · maks. 30 plików',layout:'Zdjęcia na stronę',one:'1 zdjęcie',two:'2 zdjęcia',four:'4 zdjęcia',print:'Drukuj / Zapisz jako PDF',clear:'Wyczyść',preview:'Podgląd PDF',pages:'stron',empty:'Dodaj dane kandydata lub zdjęcia, aby zobaczyć podgląd.',remove:'Usuń',footer:'Dane i zdjęcia pozostają na tym urządzeniu. Nic nie jest wysyłane na serwer.',cover:'Karta kandydata',candidate:'Kandydat',created:'Utworzono',noPhotos:'Nie wybrano zdjęć',tooMany:'Maksymalnie 30 zdjęć.',badFile:'Dozwolone są tylko pliki graficzne.',tooLarge:'Jeden plik jest za duży. Maksymalnie 15 MB.',totalLarge:'Wybrane pliki przekraczają 100 MB.'},
    ru:{brand:'Генератор PDF',back:'Анкета',title:'Создать PDF кандидата',lead:'Добавьте данные кандидата и до 30 фотографий документов. Всё остаётся на этом устройстве.',privacy:'Файлы никуда не загружаются. Нажмите «Печать» и выберите «Сохранить как PDF».',language:'Язык',firstName:'Имя',lastName:'Фамилия',phone:'Телефон',citizenship:'Гражданство',location:'Локация работы',recruiter:'Рекрутер',notes:'Примечание',photos:'Фотографии документов',upload:'Добавить фотографии',uploadHint:'JPG, PNG, WEBP или HEIC · до 30 файлов',layout:'Фотографий на странице',one:'1 фото',two:'2 фото',four:'4 фото',print:'Печать / Сохранить PDF',clear:'Очистить',preview:'Предпросмотр PDF',pages:'стр.',empty:'Добавьте данные кандидата или фотографии для предпросмотра.',remove:'Удалить',footer:'Данные и фотографии остаются на устройстве. Ничего не отправляется на сервер.',cover:'Карточка кандидата',candidate:'Кандидат',created:'Создано',noPhotos:'Фотографии не выбраны',tooMany:'Можно добавить максимум 30 фотографий.',badFile:'Разрешены только изображения.',tooLarge:'Один файл слишком большой. Максимум 15 МБ.',totalLarge:'Общий размер файлов превышает 100 МБ.'},
    uk:{brand:'Генератор PDF',back:'Анкета',title:'Створити PDF кандидата',lead:'Додайте дані кандидата та до 30 фотографій документів. Усе залишається на цьому пристрої.',privacy:'Файли нікуди не завантажуються. Натисніть «Друк» і виберіть «Зберегти як PDF».',language:'Мова',firstName:'Ім’я',lastName:'Прізвище',phone:'Телефон',citizenship:'Громадянство',location:'Локація роботи',recruiter:'Рекрутер',notes:'Примітка',photos:'Фотографії документів',upload:'Додати фотографії',uploadHint:'JPG, PNG, WEBP або HEIC · до 30 файлів',layout:'Фотографій на сторінці',one:'1 фото',two:'2 фото',four:'4 фото',print:'Друк / Зберегти PDF',clear:'Очистити',preview:'Перегляд PDF',pages:'стор.',empty:'Додайте дані кандидата або фотографії для перегляду.',remove:'Видалити',footer:'Дані та фотографії залишаються на пристрої. Нічого не надсилається на сервер.',cover:'Картка кандидата',candidate:'Кандидат',created:'Створено',noPhotos:'Фотографії не вибрано',tooMany:'Можна додати максимум 30 фотографій.',badFile:'Дозволені лише зображення.',tooLarge:'Один файл завеликий. Максимум 15 МБ.',totalLarge:'Загальний розмір файлів перевищує 100 МБ.'},
    ka:{brand:'PDF გენერატორი',back:'განაცხადი',title:'კანდიდატის PDF-ის შექმნა',lead:'დაამატეთ კანდიდატის მონაცემები და მაქსიმუმ 30 დოკუმენტის ფოტო. ყველაფერი რჩება ამ მოწყობილობაზე.',privacy:'ფაილები არ იტვირთება. აირჩიეთ ბეჭდვა და შემდეგ PDF-ად შენახვა.',language:'ენა',firstName:'სახელი',lastName:'გვარი',phone:'ტელეფონი',citizenship:'მოქალაქეობა',location:'სამუშაო ადგილი',recruiter:'რეკრუტერი',notes:'შენიშვნა',photos:'დოკუმენტის ფოტოები',upload:'ფოტოების დამატება',uploadHint:'JPG, PNG, WEBP ან HEIC · მაქს. 30',layout:'ფოტოები გვერდზე',one:'1 ფოტო',two:'2 ფოტო',four:'4 ფოტო',print:'ბეჭდვა / PDF შენახვა',clear:'გასუფთავება',preview:'PDF გადახედვა',pages:'გვერდი',empty:'გადახედვისთვის დაამატეთ მონაცემები ან ფოტოები.',remove:'წაშლა',footer:'მონაცემები და ფოტოები რჩება მოწყობილობაზე.',cover:'კანდიდატის ბარათი',candidate:'კანდიდატი',created:'შექმნილია',noPhotos:'ფოტოები არ არის',tooMany:'მაქსიმუმ 30 ფოტო.',badFile:'დაშვებულია მხოლოდ სურათები.',tooLarge:'ფაილი ძალიან დიდია. მაქს. 15 MB.',totalLarge:'ფაილების საერთო ზომა 100 MB-ს აჭარბებს.'},
    az:{brand:'PDF generatoru',back:'Ərizə',title:'Namizəd PDF-i yaradın',lead:'Namizəd məlumatlarını və 30-a qədər sənəd şəklini əlavə edin. Hər şey bu cihazda qalır.',privacy:'Fayllar yüklənmir. Çap et və PDF kimi saxla seçin.',language:'Dil',firstName:'Ad',lastName:'Soyad',phone:'Telefon',citizenship:'Vətəndaşlıq',location:'İş yeri',recruiter:'Rekruter',notes:'Qeyd',photos:'Sənəd şəkilləri',upload:'Şəkil əlavə et',uploadHint:'JPG, PNG, WEBP və ya HEIC · maksimum 30',layout:'Səhifədə şəkil',one:'1 şəkil',two:'2 şəkil',four:'4 şəkil',print:'Çap et / PDF saxla',clear:'Təmizlə',preview:'PDF önbaxış',pages:'səhifə',empty:'Önbaxış üçün məlumat və ya şəkil əlavə edin.',remove:'Sil',footer:'Məlumat və şəkillər cihazda qalır.',cover:'Namizəd kartı',candidate:'Namizəd',created:'Yaradılıb',noPhotos:'Şəkil seçilməyib',tooMany:'Maksimum 30 şəkil.',badFile:'Yalnız şəkillərə icazə verilir.',tooLarge:'Fayl çox böyükdür. Maksimum 15 MB.',totalLarge:'Ümumi ölçü 100 MB-dan çoxdur.'},
    hy:{brand:'PDF գեներատոր',back:'Հայտ',title:'Ստեղծել թեկնածուի PDF',lead:'Ավելացրեք թեկնածուի տվյալները և մինչև 30 փաստաթղթի լուսանկար։ Ամեն ինչ մնում է սարքում։',privacy:'Ֆայլերը չեն բեռնվում։ Ընտրեք Տպել և Պահել որպես PDF։',language:'Լեզու',firstName:'Անուն',lastName:'Ազգանուն',phone:'Հեռախոս',citizenship:'Քաղաքացիություն',location:'Աշխատանքի վայր',recruiter:'Հավաքագրող',notes:'Նշում',photos:'Փաստաթղթերի լուսանկարներ',upload:'Ավելացնել լուսանկարներ',uploadHint:'JPG, PNG, WEBP կամ HEIC · առավելագույնը 30',layout:'Լուսանկարներ էջում',one:'1 լուսանկար',two:'2 լուսանկար',four:'4 լուսանկար',print:'Տպել / Պահել PDF',clear:'Մաքրել',preview:'PDF դիտում',pages:'էջ',empty:'Դիտման համար ավելացրեք տվյալներ կամ լուսանկարներ։',remove:'Ջնջել',footer:'Տվյալներն ու լուսանկարները մնում են սարքում։',cover:'Թեկնածուի քարտ',candidate:'Թեկնածու',created:'Ստեղծված է',noPhotos:'Լուսանկար չկա',tooMany:'Առավելագույնը 30 լուսանկար։',badFile:'Թույլատրվում են միայն պատկերներ։',tooLarge:'Ֆայլը մեծ է։ Առավելագույնը 15 ՄԲ։',totalLarge:'Ընդհանուր չափը գերազանցում է 100 ՄԲ։'},
    tr:{brand:'PDF oluşturucu',back:'Başvuru',title:'Aday PDF’i oluştur',lead:'Aday bilgilerini ve en fazla 30 belge fotoğrafını ekleyin. Her şey bu cihazda kalır.',privacy:'Dosyalar yüklenmez. Yazdır ve PDF olarak kaydet seçin.',language:'Dil',firstName:'Ad',lastName:'Soyad',phone:'Telefon',citizenship:'Vatandaşlık',location:'İş yeri',recruiter:'İşe alım uzmanı',notes:'Not',photos:'Belge fotoğrafları',upload:'Fotoğraf ekle',uploadHint:'JPG, PNG, WEBP veya HEIC · en fazla 30',layout:'Sayfa başına fotoğraf',one:'1 fotoğraf',two:'2 fotoğraf',four:'4 fotoğraf',print:'Yazdır / PDF kaydet',clear:'Temizle',preview:'PDF önizleme',pages:'sayfa',empty:'Önizleme için veri veya fotoğraf ekleyin.',remove:'Sil',footer:'Veriler ve fotoğraflar cihazda kalır.',cover:'Aday kartı',candidate:'Aday',created:'Oluşturuldu',noPhotos:'Fotoğraf seçilmedi',tooMany:'En fazla 30 fotoğraf.',badFile:'Yalnızca görseller kabul edilir.',tooLarge:'Dosya çok büyük. En fazla 15 MB.',totalLarge:'Toplam boyut 100 MB’ı aşıyor.'},
    uz:{brand:'PDF generator',back:'Ariza',title:'Nomzod PDF yarating',lead:'Nomzod ma’lumotlari va 30 tagacha hujjat rasmini qo‘shing. Hammasi shu qurilmada qoladi.',privacy:'Fayllar yuklanmaydi. Chop etish va PDF sifatida saqlashni tanlang.',language:'Til',firstName:'Ism',lastName:'Familiya',phone:'Telefon',citizenship:'Fuqarolik',location:'Ish joyi',recruiter:'Rekruter',notes:'Izoh',photos:'Hujjat rasmlari',upload:'Rasm qo‘shish',uploadHint:'JPG, PNG, WEBP yoki HEIC · 30 tagacha',layout:'Sahifadagi rasmlar',one:'1 rasm',two:'2 rasm',four:'4 rasm',print:'Chop etish / PDF saqlash',clear:'Tozalash',preview:'PDF ko‘rinishi',pages:'sahifa',empty:'Ko‘rish uchun ma’lumot yoki rasm qo‘shing.',remove:'O‘chirish',footer:'Ma’lumot va rasmlar qurilmada qoladi.',cover:'Nomzod kartasi',candidate:'Nomzod',created:'Yaratildi',noPhotos:'Rasm tanlanmagan',tooMany:'Ko‘pi bilan 30 rasm.',badFile:'Faqat rasmlar qabul qilinadi.',tooLarge:'Fayl juda katta. Maksimum 15 MB.',totalLarge:'Umumiy hajm 100 MB dan oshdi.'}
  };
  const languageNames = MASTER.languages;
  const priority = ['pl','uk','ru','en','ka','az','hy','tr','uz'];
  const state = {lang:'pl',layout:2,firstName:'',lastName:'',phone:'',citizenship:'',location:'',recruiter:'',notes:'',files:[],message:''};
  const maxFiles = 30, maxFile = 15*1024*1024, maxTotal = 100*1024*1024;
  const esc = (value) => String(value ?? '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#039;');
  const t = (key) => (UI[state.lang] || UI.en)[key] || UI.en[key] || key;
  const formatBytes = (bytes) => bytes < 1024*1024 ? `${Math.max(1,Math.round(bytes/1024))} KB` : `${(bytes/1024/1024).toFixed(1)} MB`;
  const fullName = () => `${state.firstName} ${state.lastName}`.trim() || '—';
  const createdAt = () => new Intl.DateTimeFormat('pl-PL',{dateStyle:'medium',timeStyle:'short'}).format(new Date());

  function render(){
    const meta = languageNames[state.lang] || languageNames.en;
    document.documentElement.lang = state.lang; document.documentElement.dir = meta.direction || 'ltr';
    document.getElementById('brandSubtitle').textContent = t('brand'); document.getElementById('backLink').textContent = `← ${t('back')}`; document.getElementById('footerText').textContent = t('footer');
    app.innerHTML = `<section class="generator-grid"><div class="panel controls no-print"><span class="eyebrow">PDF · ${esc(t('brand'))}</span><h1>${esc(t('title'))}</h1><p class="lead">${esc(t('lead'))}</p><div class="privacy-note">🔒 ${esc(t('privacy'))}</div>${state.message ? `<div class="status-note error">${esc(state.message)}</div>`:''}<div class="field"><label for="pdfLanguage">${esc(t('language'))}</label><select id="pdfLanguage">${Object.entries(languageNames).sort(([a],[b])=>(priority.indexOf(a)<0?99:priority.indexOf(a))-(priority.indexOf(b)<0?99:priority.indexOf(b))).map(([code,m])=>`<option value="${code}" ${code===state.lang?'selected':''}>${m.flag} ${esc(m.native)}</option>`).join('')}</select></div><div class="field-grid">${input('firstName',t('firstName'))}${input('lastName',t('lastName'))}${input('phone',t('phone'),'tel')}${input('citizenship',t('citizenship'))}${input('location',t('location'))}${input('recruiter',t('recruiter'))}</div><div class="field"><label for="notes">${esc(t('notes'))}</label><textarea id="notes" data-field="notes" maxlength="1200">${esc(state.notes)}</textarea></div><div class="field"><span class="field-label">${esc(t('photos'))}</span><label class="upload-box"><input id="photoFiles" type="file" multiple accept="image/jpeg,image/png,image/webp,image/heic,image/heif" capture="environment"><strong>＋ ${esc(t('upload'))}</strong><small>${esc(t('uploadHint'))}</small></label><div class="file-list">${fileRows()}</div></div><div class="field"><label for="layout">${esc(t('layout'))}</label><select id="layout"><option value="1" ${state.layout===1?'selected':''}>${esc(t('one'))}</option><option value="2" ${state.layout===2?'selected':''}>${esc(t('two'))}</option><option value="4" ${state.layout===4?'selected':''}>${esc(t('four'))}</option></select></div><div class="actions"><button class="button primary" type="button" data-action="print">${esc(t('print'))}</button><button class="button secondary" type="button" data-action="clear">${esc(t('clear'))}</button></div><p class="help">${esc(t('privacy'))}</p></div><div class="panel preview-panel"><div class="preview-title"><span>${esc(t('preview'))}</span><span class="page-count">${pageCount()} ${esc(t('pages'))}</span></div><div class="pdf-pages">${previewPages()}</div></div></section>`;
    bind();
  }
  function input(id,label,type='text'){return `<div class="field"><label for="${id}">${esc(label)}</label><input id="${id}" data-field="${id}" type="${type}" value="${esc(state[id])}" maxlength="140"></div>`}
  function fileRows(){if(!state.files.length)return `<div class="status-note">${esc(t('noPhotos'))}</div>`;return state.files.map((item,index)=>`<div class="file-row"><span class="file-thumb">${item.previewable?`<img src="${item.url}" alt="">`:'HEIC'}</span><span class="file-copy"><strong>${esc(item.file.name)}</strong><small>${formatBytes(item.file.size)} · ${item.rotation}°</small></span><span class="file-actions"><button type="button" data-file-action="up" data-index="${index}" aria-label="up">↑</button><button type="button" data-file-action="down" data-index="${index}" aria-label="down">↓</button><button type="button" data-file-action="rotate" data-index="${index}" aria-label="rotate">↻</button><button class="danger" type="button" data-file-action="remove" data-index="${index}" aria-label="${esc(t('remove'))}">×</button></span></div>`).join('')}
  function pageCount(){return 1+Math.ceil(state.files.length/state.layout)}
  function previewPages(){const cover=`<article class="pdf-sheet pdf-cover"><div class="pdf-brand"><img src="../../assets/citronex-logo.jpg" alt="Citronex"><span>${esc(t('cover'))}</span></div><h2 class="pdf-title">${esc(fullName())}</h2><dl class="pdf-data"><dt>${esc(t('phone'))}</dt><dd class="ltr">${esc(state.phone||'—')}</dd><dt>${esc(t('citizenship'))}</dt><dd>${esc(state.citizenship||'—')}</dd><dt>${esc(t('location'))}</dt><dd>${esc(state.location||'—')}</dd><dt>${esc(t('recruiter'))}</dt><dd>${esc(state.recruiter||'—')}</dd></dl>${state.notes?`<div class="pdf-note">${esc(state.notes)}</div>`:''}<div class="pdf-footer"><span>${esc(t('candidate'))}</span><span>${esc(t('created'))}: ${esc(createdAt())}</span></div></article>`;if(!state.files.length)return cover;const pages=[];for(let i=0;i<state.files.length;i+=state.layout){const group=state.files.slice(i,i+state.layout);pages.push(`<article class="pdf-sheet photo-sheet photos-${state.layout}">${group.map(item=>`<figure>${item.previewable?`<img class="rot-${item.rotation}" src="${item.url}" alt="${esc(item.file.name)}">`:`<div class="empty-preview">${esc(item.file.name)}</div>`}<figcaption>${esc(item.file.name)}</figcaption></figure>`).join('')}</article>`)}return cover+pages.join('')}
  function bind(){document.getElementById('pdfLanguage').addEventListener('change',e=>{state.lang=e.target.value;state.message='';render()});app.querySelectorAll('[data-field]').forEach(el=>el.addEventListener('input',()=>{state[el.dataset.field]=el.value;renderPreviewOnly()}));document.getElementById('layout').addEventListener('change',e=>{state.layout=Number(e.target.value);renderPreviewOnly()});document.getElementById('photoFiles').addEventListener('change',e=>addFiles([...e.target.files]));app.querySelectorAll('[data-file-action]').forEach(btn=>btn.addEventListener('click',()=>fileAction(btn.dataset.fileAction,Number(btn.dataset.index))));app.querySelector('[data-action="clear"]').addEventListener('click',clearAll);app.querySelector('[data-action="print"]').addEventListener('click',()=>{window.__PDF_PRINT_REQUESTED__=true;window.dispatchEvent(new CustomEvent('citronex:pdf-print'));if(!window.__CITRONEX_TEST_MODE__)window.print()})}
  function renderPreviewOnly(){const panel=app.querySelector('.pdf-pages');const count=app.querySelector('.page-count');if(panel)panel.innerHTML=previewPages();if(count)count.textContent=`${pageCount()} ${t('pages')}`}
  function addFiles(incoming){state.message='';for(const file of incoming){if(state.files.length>=maxFiles){state.message=t('tooMany');break}if(!file.type.startsWith('image/')){state.message=t('badFile');continue}if(file.size>maxFile){state.message=t('tooLarge');continue}if(state.files.reduce((sum,item)=>sum+item.file.size,0)+file.size>maxTotal){state.message=t('totalLarge');continue}const url=URL.createObjectURL(file);state.files.push({file,url,rotation:0,previewable:!/(heic|heif)$/i.test(file.name)})}render()}
  function fileAction(action,index){const item=state.files[index];if(!item)return;if(action==='remove'){URL.revokeObjectURL(item.url);state.files.splice(index,1)}else if(action==='rotate')item.rotation=(item.rotation+90)%360;else if(action==='up'&&index>0)[state.files[index-1],state.files[index]]=[state.files[index],state.files[index-1]];else if(action==='down'&&index<state.files.length-1)[state.files[index+1],state.files[index]]=[state.files[index],state.files[index+1]];render()}
  function clearAll(){state.files.forEach(item=>URL.revokeObjectURL(item.url));Object.assign(state,{layout:2,firstName:'',lastName:'',phone:'',citizenship:'',location:'',recruiter:'',notes:'',files:[],message:''});render()}
  window.addEventListener('pagehide',()=>state.files.forEach(item=>URL.revokeObjectURL(item.url)),{once:true});render();
})();'''

BUILD_EXCEL = r'''from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter, quote_sheetname

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / 'apply' / 'excel'
OUT.mkdir(parents=True, exist_ok=True)
COLUMNS = [
'ID zgłoszenia','Data zgłoszenia','SLA do','Język','Rekruter','E-mail rekrutera','Ankietę wypełnia','Osoba / partner wypełniający','Kod grupy / partnera','Preferowana lokalizacja','Imię','Nazwisko','Telefon','Komunikator','E-mail kandydata','Obywatelstwo','Kraj pobytu','Miasto','Wiek','W Polsce','Dokument pobytowy','Stanowisko','Doświadczenie','Gotowość','Praca zmianowa','Zakwaterowanie','Źródło deklarowane','Szczegóły źródła / polecający','Źródło linku','Kampania','Wakacja / oferta','Komentarz','Status','Data pierwszego kontaktu','Liczba prób kontaktu','Następny kontakt','Wynik rozmowy','Decyzja','Powód odmowy','Uwagi rekrutera']
EXTRA = ['Status SLA','Dni w procesie','Duplikat telefonu']
RECRUITERS = ['Yana Radushynska','Yuliia Korniienko','Fariz Injaev','Oleksandr Kiris','Maksym Saliuk','Anastasiia Derepa']
FILES = [('Recruitment_Master.xlsx',None),('Recruitment_yana.xlsx',RECRUITERS[0]),('Recruitment_yuliia.xlsx',RECRUITERS[1]),('Recruitment_fariz.xlsx',RECRUITERS[2]),('Recruitment_oleksandr.xlsx',RECRUITERS[3]),('Recruitment_maksym.xlsx',RECRUITERS[4]),('Recruitment_anastasiia.xlsx',RECRUITERS[5])]
STATUSES=['NOWY','PRZYPISANY','DO KONTAKTU','NIE ODEBRAŁ','PONOWNY KONTAKT','ROZMOWA UMÓWIONA','ROZMOWA PRZEPROWADZONA','DO DECYZJI','REZERWA','ZAAKCEPTOWANY','DO LEGALIZACJI','ZATRUDNIONY','ODRZUCONY','REZYGNACJA','DUPLIKAT']
DECISIONS=['','DO DECYZJI','ZAAKCEPTOWANY','REZERWA','ODRZUCONY','REZYGNACJA']
LOCATIONS=['Siechnice','Ryczywół / Kozienice','Bogatynia','Zgorzelec','Pruszcz Gdański','Dowolna lokalizacja']
RESULTS=['','Połączono','Nie odebrał','Numer nieaktywny','Oddzwoni później','Rozmowa umówiona','Nie zainteresowany','Błędny numer']
HEAD_FILL=PatternFill('solid',fgColor='075B39'); HEAD_FONT=Font(color='FFFFFF',bold=True); THIN=Side(style='thin',color='D8E5DE'); BORDER=Border(bottom=THIN)

def dt_formula(cell):
    return f'DATE(LEFT({cell},4),MID({cell},6,2),MID({cell},9,2))+TIME(MID({cell},12,2),MID({cell},15,2),MID({cell},18,2))'

def make(filename, recruiter):
    wb=Workbook(); ws=wb.active; ws.title='KOLEJKA'; ws.freeze_panes='A2'; ws.sheet_view.showGridLines=False
    headers=COLUMNS+EXTRA
    for c,h in enumerate(headers,1):
        cell=ws.cell(1,c,h); cell.fill=HEAD_FILL; cell.font=HEAD_FONT; cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=BORDER
    ws.row_dimensions[1].height=42
    widths={1:24,2:19,3:19,4:9,5:23,6:30,7:21,8:28,9:20,10:25,11:18,12:20,13:18,14:16,15:28,16:18,17:18,18:18,19:8,20:12,21:24,22:24,23:30,24:20,25:18,26:17,27:20,28:34,29:20,30:20,31:22,32:35,33:22,34:20,35:14,36:20,37:24,38:20,39:28,40:38,41:18,42:15,43:20}
    for c,w in widths.items(): ws.column_dimensions[get_column_letter(c)].width=w
    for r in range(2,2002):
        deadline=dt_formula(f'C{r}'); submitted=dt_formula(f'B{r}')
        ws.cell(r,41,f'=IF(A{r}="","",IF(AH{r}<>"","ZREALIZOWANE",IF(NOW()>{deadline},"PRZEKROCZONE",IF({deadline}-NOW()<=0.25,"ZAGROŻONE","W TERMINIE"))))')
        ws.cell(r,42,f'=IF(A{r}="","",MAX(0,INT(NOW()-({submitted}))))')
        ws.cell(r,43,f'=IF(M{r}="","",IF(COUNTIF($M$2:$M$2001,M{r})>1,"DUPLIKAT",""))')
        for c in range(1,44): ws.cell(r,c).alignment=Alignment(vertical='top',wrap_text=True)
    ws.auto_filter.ref='A1:AQ2001'
    dictionaries=wb.create_sheet('SŁOWNIKI'); dictionaries.sheet_state='hidden'
    lists=[('Status',STATUSES),('Decyzja',DECISIONS),('Lokalizacja',LOCATIONS),('Wynik rozmowy',RESULTS),('Rekruter',RECRUITERS)]
    for col,(title,values) in enumerate(lists,1):
        dictionaries.cell(1,col,title)
        for row,value in enumerate(values,2): dictionaries.cell(row,col,value)
    def add_validation(column, formula, start=2,end=2001):
        dv=DataValidation(type='list',formula1=formula,allow_blank=True); ws.add_data_validation(dv); dv.add(f'{column}{start}:{column}{end}')
    add_validation('AG',f"{quote_sheetname('SŁOWNIKI')}!$A$2:$A${len(STATUSES)+1}")
    add_validation('AL',f"{quote_sheetname('SŁOWNIKI')}!$B$2:$B${len(DECISIONS)+1}")
    add_validation('J',f"{quote_sheetname('SŁOWNIKI')}!$C$2:$C${len(LOCATIONS)+1}")
    add_validation('AK',f"{quote_sheetname('SŁOWNIKI')}!$D$2:$D${len(RESULTS)+1}")
    add_validation('E',f"{quote_sheetname('SŁOWNIKI')}!$E$2:$E${len(RECRUITERS)+1}")
    red=PatternFill('solid',fgColor='FECACA'); amber=PatternFill('solid',fgColor='FEF3C7'); green=PatternFill('solid',fgColor='DCFCE7'); blue=PatternFill('solid',fgColor='DBEAFE')
    ws.conditional_formatting.add('AO2:AO2001',FormulaRule(formula=['AO2="PRZEKROCZONE"'],fill=red))
    ws.conditional_formatting.add('AO2:AO2001',FormulaRule(formula=['AO2="ZAGROŻONE"'],fill=amber))
    ws.conditional_formatting.add('AO2:AO2001',FormulaRule(formula=['AO2="W TERMINIE"'],fill=green))
    ws.conditional_formatting.add('AO2:AO2001',FormulaRule(formula=['AO2="ZREALIZOWANE"'],fill=blue))
    ws.conditional_formatting.add('AQ2:AQ2001',FormulaRule(formula=['AQ2="DUPLIKAT"'],fill=red))
    dash=wb.create_sheet('PULPIT'); dash.sheet_view.showGridLines=False
    dash['A1']='CITRONEX / PPO SIECHNICE — KOLEJKA REKRUTACYJNA'; dash['A1'].font=Font(size=18,bold=True,color='075B39'); dash.merge_cells('A1:D1')
    dash['A3']='Zakres'; dash['B3']=recruiter or 'WSZYSCY REKRUTERZY'; dash['B3'].font=Font(bold=True)
    recruiter_condition='' if recruiter is None else f',KOLEJKA!$E$2:$E$2001,"{recruiter}"'
    metrics=[('Wszystkie zgłoszenia',f'=COUNTIF(KOLEJKA!$A$2:$A$2001,"<>"{recruiter_condition})'),('Nowe',f'=COUNTIFS(KOLEJKA!$AG$2:$AG$2001,"NOWY"{recruiter_condition})'),('SLA przekroczone',f'=COUNTIFS(KOLEJKA!$AO$2:$AO$2001,"PRZEKROCZONE"{recruiter_condition})'),('Do kontaktu',f'=COUNTIFS(KOLEJKA!$AG$2:$AG$2001,"DO KONTAKTU"{recruiter_condition})'),('Zaakceptowani',f'=COUNTIFS(KOLEJKA!$AG$2:$AG$2001,"ZAAKCEPTOWANY"{recruiter_condition})'),('Zatrudnieni',f'=COUNTIFS(KOLEJKA!$AG$2:$AG$2001,"ZATRUDNIONY"{recruiter_condition})')]
    for r,(label,formula) in enumerate(metrics,5): dash.cell(r,1,label); dash.cell(r,2,formula); dash.cell(r,2).font=Font(size=16,bold=True,color='075B39')
    dash.column_dimensions['A'].width=30; dash.column_dimensions['B'].width=24
    ins=wb.create_sheet('INSTRUKCJA'); ins.sheet_view.showGridLines=False; ins['A1']='Jak używać pliku'; ins['A1'].font=Font(size=18,bold=True,color='075B39')
    lines=['1. Otwórz wiadomość od kandydata.','2. Skopiuj wyłącznie jeden wiersz TSV znajdujący się pod napisem DANE DO EXCEL.','3. W arkuszu KOLEJKA wybierz pierwszą pustą komórkę w kolumnie A.','4. Wklej dane — pierwsze 40 kolumn zostanie uzupełnionych automatycznie.','5. Zmień Status, uzupełnij pierwszy kontakt, następny kontakt, wynik rozmowy i decyzję.','6. Kolumny Status SLA, Dni w procesie i Duplikat telefonu obliczają się automatycznie.']
    for r,line in enumerate(lines,3): ins.cell(r,1,line); ins.cell(r,1).alignment=Alignment(wrap_text=True); ins.row_dimensions[r].height=28
    ins.column_dimensions['A'].width=110
    wb.save(OUT/filename)

for item in FILES: make(*item)
print('Generated',len(FILES),'recruitment workbooks.')
'''

VALIDATE_EXCEL = r'''from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET
ROOT=Path(__file__).resolve().parents[1]; OUT=ROOT/'apply'/'excel'
FILES=['Recruitment_Master.xlsx','Recruitment_yana.xlsx','Recruitment_yuliia.xlsx','Recruitment_fariz.xlsx','Recruitment_oleksandr.xlsx','Recruitment_maksym.xlsx','Recruitment_anastasiia.xlsx']
HEADERS=['ID zgłoszenia','Data zgłoszenia','SLA do','Język','Rekruter','E-mail rekrutera','Ankietę wypełnia','Osoba / partner wypełniający','Kod grupy / partnera','Preferowana lokalizacja','Imię','Nazwisko','Telefon','Komunikator','E-mail kandydata','Obywatelstwo','Kraj pobytu','Miasto','Wiek','W Polsce','Dokument pobytowy','Stanowisko','Doświadczenie','Gotowość','Praca zmianowa','Zakwaterowanie','Źródło deklarowane','Szczegóły źródła / polecający','Źródło linku','Kampania','Wakacja / oferta','Komentarz','Status','Data pierwszego kontaktu','Liczba prób kontaktu','Następny kontakt','Wynik rozmowy','Decyzja','Powód odmowy','Uwagi rekrutera','Status SLA','Dni w procesie','Duplikat telefonu']
NS={'m':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
def value(cell,shared):
    t=cell.attrib.get('t'); v=cell.find('m:v',NS)
    if t=='inlineStr':
        node=cell.find('m:is/m:t',NS); return node.text if node is not None else ''
    if v is None:return ''
    if t=='s':return shared[int(v.text)]
    return v.text or ''
for name in FILES:
    path=OUT/name
    assert path.exists() and path.stat().st_size>10000, f'missing or too small: {name}'
    with zipfile.ZipFile(path) as z:
        assert z.testzip() is None, f'corrupt zip: {name}'
        shared=[]
        if 'xl/sharedStrings.xml' in z.namelist():
            root=ET.fromstring(z.read('xl/sharedStrings.xml'))
            shared=[''.join(n.text or '' for n in si.findall('.//m:t',NS)) for si in root.findall('m:si',NS)]
        sheet=ET.fromstring(z.read('xl/worksheets/sheet1.xml'))
        row=sheet.find('.//m:sheetData/m:row[@r="1"]',NS)
        values=[value(c,shared) for c in row.findall('m:c',NS)]
        assert values[:43]==HEADERS, f'header mismatch in {name}: {values[:5]}'
        formulas=[n.text or '' for n in sheet.findall('.//m:f',NS)]
        assert any('COUNTIF' in f for f in formulas), f'duplicate formula missing in {name}'
        assert any('NOW()' in f for f in formulas), f'SLA formula missing in {name}'
print('Excel validation passed:',len(FILES),'files, 40 input columns + 3 automatic columns.')
'''

VALIDATE_APPLY = r'''import fs from 'node:fs';import vm from 'node:vm';
const read=p=>fs.readFileSync(p,'utf8');const fail=m=>{console.error(`Application validation failed: ${m}`);process.exit(1)};const assert=(c,m)=>{if(!c)fail(m)};
const context=vm.createContext({window:{},console,URL,Intl,Object});for(const p of ['apply/config.js','apply/translations.js'])vm.runInContext(read(p),context,{filename:p});
const config=context.window.CITRONEX_SIMPLE_CONFIG,i18n=context.window.CITRONEX_SIMPLE_I18N;assert(config&&i18n,'configuration or translations missing');assert(config.version==='17.0.0','wrong version');assert(config.recruiters.length===6,'six recruiters required');assert(config.locations.length===6,'six locations required');assert(config.excelColumns.length===40,'40 Excel columns required');assert(Object.keys(i18n.languages).length===20,'20 languages required');
for(const code of Object.keys(i18n.languages)){const l=i18n.locales[code];assert(l,`missing locale ${code}`);for(const key of ['chooseLanguage','chooseRecruiter','step1Title','step2Title','step3Title','step4Title','sendRow','options'])assert(l[key],`missing ${code}.${key}`);for(const group of ['messenger','yesNo','documents','jobs','starts','shifts','sources','locations'])assert(l.options?.[group]&&Object.keys(l.options[group]).length,`missing ${code}.options.${group}`)}
const index=read('apply/index.html'),app=read('apply/app.js'),css=read('apply/styles.css');for(const m of ['styles.css?v=17.0.0','config.js?v=17.0.0','translations.js?v=17.0.0','app.js?v=17.0.0','href="pdf/"','../assets/citronex-logo.jpg'])assert(index.includes(m),`index missing ${m}`);assert((index.match(/rel="stylesheet"/g)||[]).length===1,'main form must load one stylesheet');for(const old of ['i18n-core.js','i18n-caucasus-central.js','i18n-asia.js','i18n-extra.js','i18n-mobile.js','mobile-v8','ui-v9','ui-v11','delivery-v10','app-mobile'])assert(!index.includes(old),`old layer loaded: ${old}`);
for(const m of ['function rowValues','function sendRow','window.__CITRONEX_LAST_MAILTO__','window.__CITRONEX_TEST_MODE__','locale().options?.[group]','closest(\'.choice-card\')?.classList.toggle(\'selected\'','DANE DO EXCEL','window.location.assign(mailto)'])assert(app.includes(m),`app missing ${m}`);for(const forbidden of ['type="file"','navigator.share','new File(','buildPdf','TABELA KANDYDATA'])assert(!app.includes(forbidden),`main form contains removed workflow: ${forbidden}`);assert(!/position\s*:\s*(sticky|fixed)/i.test(css),'main CSS contains sticky/fixed positioning');assert(!css.includes('backdrop-filter'),'main CSS contains backdrop filter');for(const m of ['@media(max-width:700px)','@media(max-width:440px)','.form-grid','.choice-card','.send-box'])assert(css.includes(m),`CSS missing ${m}`);
console.log(`Application validation passed: clean v${config.version}, ${Object.keys(i18n.languages).length} languages, ${config.recruiters.length} recruiters, ${config.excelColumns.length} columns.`);
'''

VALIDATE_DELIVERY = r'''import fs from 'node:fs';const read=p=>fs.readFileSync(p,'utf8');const fail=m=>{console.error(`PDF validation failed: ${m}`);process.exit(1)};const assert=(c,m)=>{if(!c)fail(m)};
const main=read('apply/index.html'),pdf=read('apply/pdf/index.html'),css=read('apply/pdf/styles.css'),app=read('apply/pdf/app.js'),redirect=read('apply/offline.html');assert(main.includes('href="pdf/"'),'main form must link to PDF generator');assert(!main.includes('type="file"'),'main form contains file input');for(const m of ['styles.css?v=2.0.0','../translations.js?v=17.0.0','app.js?v=2.0.0','id="pdfApp"','../../assets/citronex-logo.jpg'])assert(pdf.includes(m),`PDF index missing ${m}`);for(const m of ['window.print()','URL.createObjectURL','type="file"','maxFiles = 30','data-file-action="rotate"','window.__PDF_PRINT_REQUESTED__','photos-2','photos-4'])assert(app.includes(m),`PDF app missing ${m}`);for(const m of ['@page{size:A4','@media print','.generator-grid{display:grid;grid-template-columns:minmax(0,420px) minmax(0,1fr)','.rot-90','.photo-sheet.photos-4'])assert(css.includes(m),`PDF CSS missing ${m}`);assert(!/position\s*:\s*(sticky|fixed)/i.test(css),'PDF CSS contains sticky/fixed positioning');for(const forbidden of [/\bcsv\b/i,/\bxlsx\b/i,/\btsv\b/i,/\bzip\b/i,/navigator\.share/,/mailto:/i,/wa\.me/i,/viber:\/\//i,/fetch\s*\(/,/serviceWorker\.register/])assert(!forbidden.test(`${pdf}\n${css}\n${app}`),`forbidden PDF workflow ${forbidden}`);assert(redirect.includes('url=pdf/'),'legacy redirect missing');console.log('PDF generator validation passed: separate print-only tool with up to 30 photos.');
'''

E2E = r'''import http from 'node:http';import fs from 'node:fs';import path from 'node:path';import { chromium } from 'playwright';
const root=process.cwd(),port=4173,base=`http://127.0.0.1:${port}`;const mime={'.html':'text/html; charset=utf-8','.js':'text/javascript; charset=utf-8','.css':'text/css; charset=utf-8','.jpg':'image/jpeg','.png':'image/png','.svg':'image/svg+xml','.xlsx':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'};
const server=http.createServer((req,res)=>{const url=new URL(req.url,base);let file=decodeURIComponent(url.pathname);if(file.endsWith('/'))file+='index.html';const full=path.resolve(root,`.${file}`);if(!full.startsWith(root)){res.writeHead(403);return res.end()}fs.readFile(full,(err,data)=>{if(err){res.writeHead(404);res.end('not found');return}res.writeHead(200,{'Content-Type':mime[path.extname(full)]||'application/octet-stream','Cache-Control':'no-store'});res.end(data)})});await new Promise(r=>server.listen(port,'127.0.0.1',r));
const browser=await chromium.launch({headless:true});const errors=[];const viewports=[{width:320,height:568},{width:360,height:800},{width:390,height:844},{width:430,height:932},{width:768,height:1024}];
async function newPage(viewport){const p=await browser.newPage({viewport});await p.addInitScript(()=>{window.__CITRONEX_TEST_MODE__=true});p.on('pageerror',e=>errors.push(e.message));p.on('console',m=>{if(m.type()==='error')errors.push(m.text())});return p}
async function choose(page,lang='ru',rec='oleksandr'){await page.goto(`${base}/apply/?e2e=${Date.now()}`,{waitUntil:'networkidle'});if(!(await page.locator(`[data-language="${lang}"]`).count()))await page.locator('[data-action="toggle-languages"]').click();await page.locator(`[data-language="${lang}"]`).click();await page.locator(`[data-recruiter="${rec}"]`).click()}
async function selectFirst(page,field){await page.locator(`select[data-field="${field}"]`).selectOption({index:1})}
async function fill(page,{partner=false}={}){if(partner){await page.locator('input[data-field="filledBy"][value="representative"]').check({force:true});await page.locator('[data-field="representativeName"]').fill('Giorgi Partner');await page.locator('[data-field="groupCode"]').fill('GE-TEST')}else await page.locator('input[data-field="filledBy"][value="self"]').check({force:true});await page.locator('[data-field="firstName"]').fill('Ivan');await page.locator('[data-field="lastName"]').fill('Petrenko');await page.locator('[data-field="phone"]').fill('+380501112233');await selectFirst(page,'messenger');await page.locator('[data-action="next"]').click();await page.locator('[data-field="citizenship"]').fill('Ukraina');await page.locator('[data-field="country"]').fill('Polska');await page.locator('[data-field="city"]').fill('Wrocław');await page.locator('[data-field="age"]').fill('32');await page.locator('input[data-field="inPoland"][value="yes"]').check({force:true});await selectFirst(page,'documents');await page.locator('[data-action="next"]').click();await selectFirst(page,'location');await selectFirst(page,'job');await selectFirst(page,'start');await page.locator('input[data-field="shift"][value="yes"]').check({force:true});await page.locator('input[data-field="housing"][value="yes"]').check({force:true});await page.locator('[data-action="next"]').click();await page.locator('input[data-field="source"][value="facebook"]').check({force:true});await page.locator('[data-field="sourceDetails"]').fill('Facebook test campaign');await page.locator('[data-field="consent"]').check();await page.locator('[data-action="next"]').click();await page.locator('[data-action="send"]').click();const mailto=await page.evaluate(()=>window.__CITRONEX_LAST_MAILTO__);if(!mailto?.startsWith('mailto:'))throw new Error('mailto was not generated');const parsed=new URL(mailto);const row=decodeURIComponent(parsed.searchParams.get('body')||'').trim().split('\n').at(-1);if(row.split('\t').length!==40)throw new Error(`TSV has ${row.split('\t').length} columns`);return mailto}
async function layoutCheck(page){const result=await page.evaluate(()=>({overflow:document.documentElement.scrollWidth>window.innerWidth+1,layered:[...document.querySelectorAll('*')].filter(el=>{const s=getComputedStyle(el);return s.position==='fixed'||s.position==='sticky'}).length}));if(result.overflow||result.layered)throw new Error(`layout failed ${JSON.stringify(result)}`)}
for(const viewport of viewports){const page=await newPage(viewport);await choose(page);await fill(page);await layoutCheck(page);await page.close()}
for(const code of Object.keys((await (await fetch(`${base}/apply/translations.js`)).text(),{}))){void code}
const languageCodes=['pl','uk','ru','en','ka','az','hy','tr','uz','ky','tg','kk','hi','bn','ne','ur','si','fil','id','vi'];for(const code of languageCodes){const page=await newPage({width:390,height:844});await page.goto(`${base}/apply/`);if(!(await page.locator(`[data-language="${code}"]`).count()))await page.locator('[data-action="toggle-languages"]').click();await page.locator(`[data-language="${code}"]`).click();if(await page.locator('[data-recruiter]').count()!==6)throw new Error(`recruiters missing for ${code}`);if(code==='ur'&&(await page.locator('html').getAttribute('dir'))!=='rtl')throw new Error('Urdu is not RTL');await layoutCheck(page);await page.close()}
const recruiterIds=['yana','yuliia','fariz','oleksandr','maksym','anastasiia'];for(const rec of recruiterIds){const page=await newPage({width:390,height:844});await choose(page,'en',rec);const mailto=await fill(page);const expected=(await page.evaluate(id=>window.CITRONEX_SIMPLE_CONFIG.recruiters.find(p=>p.id===id).email,rec));if(!mailto.startsWith(`mailto:${encodeURIComponent(expected)}`)&&!mailto.startsWith(`mailto:${expected}`))throw new Error(`wrong recipient for ${rec}`);await page.close()}
{const page=await newPage({width:390,height:844});await choose(page,'ka','oleksandr');await fill(page,{partner:true});await page.close()}
{const page=await newPage({width:390,height:844});await page.goto(`${base}/apply/pdf/`);await page.locator('[data-field="firstName"]').fill('Ivan');await page.locator('[data-field="lastName"]').fill('Petrenko');const png=Buffer.from('iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Wl2n7sAAAAASUVORK5CYII=','base64');await page.locator('#photoFiles').setInputFiles([{name:'passport.png',mimeType:'image/png',buffer:png},{name:'visa.png',mimeType:'image/png',buffer:png},{name:'cv.png',mimeType:'image/png',buffer:png}]);if(await page.locator('.file-row').count()!==3)throw new Error('PDF files were not added');await page.locator('[data-file-action="rotate"]').first().click();await page.locator('#layout').selectOption('2');if(await page.locator('.photo-sheet').count()<2)throw new Error('PDF preview pages missing');await page.locator('[data-action="print"]').click();if(!(await page.evaluate(()=>window.__PDF_PRINT_REQUESTED__)))throw new Error('print action missing');await layoutCheck(page);await page.close()}
await browser.close();await new Promise(r=>server.close(r));if(errors.length)throw new Error(`Browser errors: ${errors.join(' | ')}`);console.log('E2E passed: 5 viewports, 20 languages, 6 recruiters, partner mode, 40-column mailto and PDF generator.');
'''

PAGES = r'''name: Проверка и публикация портала
on:
  push:
    branches: ["main"]
  workflow_dispatch:
permissions:
  contents: read
  pages: write
  id-token: write
concurrency:
  group: pages
  cancel-in-progress: true
jobs:
  verify:
    runs-on: ubuntu-latest
    timeout-minutes: 20
    steps:
      - name: Получить код
        uses: actions/checkout@v7
      - name: Настроить Node.js
        uses: actions/setup-node@v7
        with:
          node-version: 22
      - name: Проверить код, анкету, PDF и Excel
        run: npm test
      - name: Установить Chromium для мобильного теста
        run: |
          npm install --no-save --ignore-scripts playwright@1.55.0
          npx playwright install --with-deps chromium
      - name: Пройти реальные браузерные сценарии
        run: node scripts/e2e-apply.mjs
      - name: Проверить чистую production-структуру
        shell: bash
        run: |
          set -euo pipefail
          test "$(grep -c 'rel="stylesheet"' apply/index.html)" -eq 1
          test "$(grep -c 'rel="stylesheet"' apply/pdf/index.html)" -eq 1
          ! grep -Eqi 'position\s*:\s*(sticky|fixed)|backdrop-filter' apply/styles.css apply/pdf/styles.css
          ! grep -Eqi 'mobile-v8|ui-v9|ui-v11|delivery-v10|app-mobile|offline-candidate' apply/index.html
          grep -q 'styles.css?v=17.0.0' apply/index.html
          grep -q 'app.js?v=17.0.0' apply/index.html
          grep -q 'styles.css?v=2.0.0' apply/pdf/index.html
          grep -q 'app.js?v=2.0.0' apply/pdf/index.html
          for f in apply/excel/Recruitment_Master.xlsx apply/excel/Recruitment_yana.xlsx apply/excel/Recruitment_yuliia.xlsx apply/excel/Recruitment_fariz.xlsx apply/excel/Recruitment_oleksandr.xlsx apply/excel/Recruitment_maksym.xlsx apply/excel/Recruitment_anastasiia.xlsx; do unzip -t "$f" >/dev/null; done
  deploy:
    environment:
      name: github-pages
      url: ${{ steps.deployment.outputs.page_url }}
    runs-on: ubuntu-latest
    timeout-minutes: 10
    needs: verify
    steps:
      - name: Получить код
        uses: actions/checkout@v7
      - name: Настроить GitHub Pages
        uses: actions/configure-pages@v6
      - name: Подготовить статические файлы
        uses: actions/upload-pages-artifact@v5
        with:
          path: "."
      - name: Опубликовать
        id: deployment
        uses: actions/deploy-pages@v5
'''

README = r'''# Citronex / PPO Siechnice — rekrutacja

## Produkcja

- Ankieta mobilna: `https://oleksandrkiris.github.io/kiris-jobs/apply/`
- Opcjonalny generator PDF: `https://oleksandrkiris.github.io/kiris-jobs/apply/pdf/`
- Excel: katalog `/apply/excel/`

## Główna ankieta

Kandydat wybiera język i rekrutera, podaje dane, lokalizację oraz źródło, sprawdza zgłoszenie i otwiera przygotowany e-mail. Wiadomość zawiera jedną linię TSV z 40 kolumnami do wklejenia w pierwszą pustą komórkę A wspólnego Excela.

Główna ankieta nie przyjmuje plików, nie tworzy PDF i nie wysyła danych do serwera. Wersja robocza pozostaje lokalnie w przeglądarce maksymalnie 24 godziny.

## Generator PDF

Generator jest całkowicie oddzielny. Pozwala wprowadzić dane, dodać do 30 zdjęć, zmienić kolejność, obracać, usuwać, wybrać 1/2/4 zdjęcia na stronie oraz zapisać wynik przez `Drukuj → Zapisz jako PDF`. Zdjęcia pozostają na urządzeniu.

## Excel

Pierwsze 40 kolumn każdego pliku odpowiada dokładnie kolejności danych w wiadomości. Dodatkowe kolumny automatyczne: `Status SLA`, `Dni w procesie`, `Duplikat telefonu`.

## Kontrola jakości

`npm test` sprawdza kod, 20 języków, 6 rekruterów, 40 kolumn, oddzielenie generatora i integralność siedmiu plików Excel. GitHub Pages dodatkowo wykonuje test Chromium na pięciu rozmiarach ekranu, wszystkich językach i wszystkich rekruterach.
'''

write(APPLY / 'config.js', CONFIG)
write(APPLY / 'index.html', INDEX)
write(APPLY / 'styles.css', STYLES)
write(PDF / 'index.html', PDF_INDEX)
write(PDF / 'styles.css', PDF_STYLES)
write(PDF / 'app.js', PDF_APP)
write(APPLY / 'offline.html', '<!doctype html><html><head><meta charset="utf-8"><meta http-equiv="refresh" content="0; url=pdf/"><title>PDF</title></head><body><a href="pdf/">Generator PDF</a></body></html>')
write(SCRIPTS / 'build-recruitment-excel.py', BUILD_EXCEL)
write(SCRIPTS / 'validate-recruitment-excel.py', VALIDATE_EXCEL)
write(SCRIPTS / 'validate-apply.mjs', VALIDATE_APPLY)
write(SCRIPTS / 'validate-delivery.mjs', VALIDATE_DELIVERY)
write(SCRIPTS / 'e2e-apply.mjs', E2E)
write(ROOT / '.github' / 'workflows' / 'pages.yml', PAGES)
write(APPLY / 'README.md', README)

# Patch the current, already complete form controller instead of layering another script over it.
app_path = APPLY / 'app.js'
app_text = app_path.read_text(encoding='utf-8')
replacements = {
    "const options = locale().options[group] || I18N.locales.en.options[group];": "const options = locale().options?.[group] || I18N.locales.en.options?.[group] || I18N.internal?.[group] || {};",
    "const options = locale().options;": "const options = { ...(I18N.locales.en.options || {}), ...(locale().options || {}) };",
    "return groups[field] ? locale().options[groups[field]]?.[value] || value : value;": "return groups[field] ? (locale().options?.[groups[field]]?.[value] || I18N.locales.en.options?.[groups[field]]?.[value] || I18N.internal?.[groups[field]]?.[value] || value) : value;"
}
for old, new in replacements.items():
    app_text = app_text.replace(old, new)
old_listener = """        delete errors[field];
        saveDraft();
        if (field === 'filledBy') renderForm();
"""
new_listener = """        delete errors[field];
        if (element.type === 'radio') {
          app.querySelectorAll(`input[name="${field}"]`).forEach((input) => {
            input.closest('.choice-card')?.classList.toggle('selected', input.checked);
          });
        }
        saveDraft();
        if (field === 'filledBy') renderForm();
"""
app_text = app_text.replace(old_listener, new_listener)
old_send = """    saveDraft();
    window.setTimeout(() => { sending = false; button.disabled = false; }, 2500);
    window.location.assign(mailto);
"""
new_send = """    saveDraft();
    window.__CITRONEX_LAST_MAILTO__ = mailto;
    window.dispatchEvent(new CustomEvent('citronex:mailto', { detail: { mailto } }));
    window.setTimeout(() => { sending = false; button.disabled = false; }, 2500);
    if (!window.__CITRONEX_TEST_MODE__) window.location.assign(mailto);
"""
app_text = app_text.replace(old_send, new_send)
required_markers = ["locale().options?.[group]", "closest('.choice-card')?.classList.toggle('selected'", "window.__CITRONEX_LAST_MAILTO__"]
for marker in required_markers:
    if marker not in app_text:
        raise SystemExit(f'Could not apply main form fix: {marker}')
write(app_path, app_text)

# Build one self-contained translations file from the existing 20-language sources.
builder_path = APPLY / 'translations.js'
builder = builder_path.read_text(encoding='utf-8')
builder = builder.replace(
    "locations: core.options?.locations || I.locales.en.options.locations",
    "locations: core.options?.locations || I.locales.en.options?.locations || E.locales?.[code]?.options?.locations || E.locales?.en?.options?.locations || {siechnice:'Siechnice',ryczywol:'Ryczywół / Kozienice',bogatynia:'Bogatynia',zgorzelec:'Zgorzelec',pruszcz:'Pruszcz Gdański',any:'Dowolna lokalizacja'}"
)
builder = builder.replace("version:'16.0.0'", "version:'17.0.0'")
tmp_builder = APPLY / '.translations-builder.js'
write(tmp_builder, builder)
node_code = r"""
const fs=require('fs'),vm=require('vm');
const ctx=vm.createContext({window:{},console,URL,Intl,Object});
for(const p of ['apply/i18n-core.js','apply/i18n-caucasus-central.js','apply/i18n-asia.js','apply/i18n-extra.js','apply/i18n-mobile.js','apply/.translations-builder.js']) vm.runInContext(fs.readFileSync(p,'utf8'),ctx,{filename:p});
const data=ctx.window.CITRONEX_SIMPLE_I18N;
if(!data||Object.keys(data.languages||{}).length!==20) throw new Error('Could not consolidate 20 languages');
for(const code of Object.keys(data.languages)){
  data.locales[code].options=data.locales[code].options||{};
  const en=data.locales.en.options||{}, internal=data.internal||{};
  for(const group of ['messenger','yesNo','documents','jobs','starts','shifts','sources','locations']) data.locales[code].options[group]=data.locales[code].options[group]||en[group]||internal[group]||{};
}
fs.writeFileSync('apply/translations.js',`(() => { 'use strict'; window.CITRONEX_SIMPLE_I18N = ${JSON.stringify(data)}; })();\n`,'utf8');
"""
subprocess.run(['node', '-e', node_code], cwd=ROOT, check=True)
tmp_builder.unlink(missing_ok=True)

# Remove obsolete translation/UI layers. They are no longer referenced.
for name in [
    'i18n-core.js','i18n-caucasus-central.js','i18n-asia.js','i18n-extra.js','i18n-mobile.js',
    'mobile-v8.css','ui-v9.css','ui-v9.js','ui-v11.css','ui-v11.js','delivery-v10.css','delivery-v10-i18n.js','delivery-v10.js',
    'app-mobile.js','offline-redirect.js','offline-v12.css','offline-shared-v14.0.3.js','offline-translations-v14.0.4.js','offline-candidate-v14.0.4.js','offline-icon.svg'
]:
    (APPLY / name).unlink(missing_ok=True)
for name in ['manifest.webmanifest','service-worker.js','app-live-v2.js']:
    (PDF / name).unlink(missing_ok=True)

# Keep package scripts aligned with the clean structure.
package_path = ROOT / 'package.json'
package = json.loads(package_path.read_text(encoding='utf-8'))
package['description'] = 'Публичная мобильная анкета Citronex / PPO Siechnice и отдельный генератор PDF'
package['scripts']['check:js'] = 'node --check assets/candidate.js && node --check assets/i18n.js && node --check assets/application-form.js && node --check apply/config.js && node --check apply/translations.js && node --check apply/app.js && node --check apply/pdf/app.js && node --check sw.js && node --check data/content.js'
package['scripts']['check:excel'] = 'python3 scripts/validate-recruitment-excel.py'
package['scripts']['test'] = 'npm run build:css && npm run generate:vacancies && npm run check:js && npm run check:content && npm run check:locales && npm run check:ui && npm run check:apply && npm run check:delivery && npm run check:excel'
package_path.write_text(json.dumps(package, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

print('Clean v17 sources prepared.')
